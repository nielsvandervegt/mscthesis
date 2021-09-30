# Pattern-Placed Revetment Finite Element Model Generator
# Script by Niels van der Vegt

# Import
import numpy as np;
import openturns as ot;
import openpyxl as op;
import time;
from abaqus import * 
from abaqusConstants import *

########################################
# MODEL SETUP AND BOUNDARY CONDITIONS
########################################

# It is possible to use distributions in the model! E.g.: ot.Normal(8,2).getSample(1)[0][0]

# Model
m_Bs = 3         # Model width in sets
m_L = 5          # Length of the 'flume bottom' in meters
g = 9.81		 # Gravitational Constant in m/s^2
relaxtime = 2.0  # Time before the first wave impact in seconds
impacttime = 2.0 # Seconds between two wave impacts
dx = 0.01		 # Numerical step in space
dt = 0.01		 # Numerical step in time

# Settings
generateLoading = True; # Set to False if you dont want generated loading (e.g. when performing pull-out tests)

# Not in use
b = 0.00        # Filter layer thickness (only geometry purposes, leave at 0.0)
i_fr = 0.00     # Fraction joint filling over top layer thickness
rhoi = 2300     # Density joint filling
ni = 0.4        # Porosity joint filling
t_w = 0.00      # Deformation of the toe in meters


########################################
# Step 1. Read the data on which samples we should make from the Excel file
########################################

# Read the 
wb = op.load_workbook(filename = 'H:\Desktop\Model\sample2.xlsx');
ws = wb.active;

# Count the amount of samples in the Excel file
check = False;
samples = 0;
while(check == False):
    if(ws.cell(row=samples+2, column=1).value == None):
        break;
    samples = samples + 1;

# Create for each input parameter an array
a_mname = []
a_slope = []
a_zdtop = []
a_zdbot = []
bb_Hs = []
a_s0p = []
a_SI = []
a_D = []
a_d = []
a_ll = []
a_kk = []
a_rhow = []
a_rhos = []
a_Fblockfilter = []
a_Fblockblock = []
aa_Hs = []
a_order = []
a_S_mid = []
a_S_width = []
a_S_amp = []
a_G_z = []
a_G_x = []

# Read the excel file into the arrays
for i in range(2,samples+2):
    a_mname.append(ws.cell(column=1, row=i).value)
    a_slope.append(float(ws.cell(column=2, row=i).value))
    a_zdtop.append(float(ws.cell(column=3, row=i).value))
    a_zdbot.append(float(ws.cell(column=4, row=i).value))
    bb_Hs.append(float(ws.cell(column=5, row=i).value))
    a_s0p.append(float(ws.cell(column=6, row=i).value))
    a_SI.append(float(ws.cell(column=7, row=i).value))
    a_D.append(float(ws.cell(column=8, row=i).value))
    a_d.append(float(ws.cell(column=9, row=i).value))
    a_ll.append(float(ws.cell(column=10, row=i).value))
    a_kk.append(float(ws.cell(column=11, row=i).value))
    a_rhow.append(float(ws.cell(column=12, row=i).value))
    a_rhos.append(float(ws.cell(column=13, row=i).value))
    a_Fblockfilter.append(float(ws.cell(column=14, row=i).value))
    a_Fblockblock.append(float(ws.cell(column=15, row=i).value))
    aa_Hs.append([float(ws.cell(column=16, row=i).value), float(ws.cell(column=17, row=i).value), float(ws.cell(column=18, row=i).value), float(ws.cell(column=19, row=i).value), float(ws.cell(column=20, row=i).value)])
    a_order.append(float(ws.cell(column=21, row=i).value))
    a_S_mid.append(float(ws.cell(column=22, row=i).value))
    a_S_width.append(float(ws.cell(column=23, row=i).value))
    a_S_amp.append(float(ws.cell(column=24, row=i).value))
    a_G_z.append(float(ws.cell(column=25, row=i).value))
    a_G_x.append(float(ws.cell(column=26, row=i).value))
wb.close;


########################################
# Step 2. Read the data on the shapes of the elements and joint filling
########################################

# Gives: length of block, center of rotation and area of each element
ginfo = np.loadtxt('Basalton\STS info.txt', delimiter=',', skiprows=1)
info = {}
for i in np.arange(len(ginfo)):
    bid = int(ginfo[i][0])
    info[bid] = {}
    info[bid]['length'] = ginfo[i][1];
    info[bid]['relrot'] = ginfo[i][2];
    info[bid]['area'] = ginfo[i][3];
    
# Same for joint filling
ginfo = np.loadtxt('Inwas\Inwasinfo.txt', delimiter=',', skiprows=1)
inwasinfo = {}
for i in np.arange(len(ginfo)):
    bid = int(ginfo[i][0])
    inwasinfo[bid] = {}
    inwasinfo[bid]['length'] = ginfo[i][1];
    inwasinfo[bid]['relrot'] = ginfo[i][2];
    inwasinfo[bid]['area'] = ginfo[i][3];



########################################
# MODEL GENERATION
########################################

start = time.time();
# Accidentally I worked with another dx for the S-profile, so I need an extra dx which will not be overwritten :)
dxx = dx 

# Create the models one by one
for i in range(len(a_mname)):
    ########################################
	# Step 3. Define the parameters for model i
	########################################
	
    modelstart = time.time();
    mname = str(a_mname[i]); # model name
    d = a_d[i]; # water depth
    rhow = a_rhow[i]; # density water
    slope = a_slope[i]; # slope
    zdtop = a_zdtop[i]; # dimensionless top of the revetment
    zdbot = a_zdbot[i]; # dimensionless bottom of the revetment
    D = a_D[i]; # top layer thickness
    rhos = a_rhos[i]; # density of the elements
    b_Hs = bb_Hs[i]; # Significant wave height
    Fblockfilter = a_Fblockfilter[i]; # Friction coefficient with filter
    Fblockblock = a_Fblockblock[i]; # Friction coefficient with block
    s0p = a_s0p[i]; # Wave steepness
    a_Hs = aa_Hs[i]; # Individual wave heights
    ll = a_ll[i]; # Leakage length
    kk = a_kk[i]; # Filter layer permability over top layer permeability
    df = ll**2 / (kk * D) # Calculate the filter thickness
    xiHs = (1/slope) / np.sqrt(s0p); # Iribarren number
    xiHs = xiHs * 0.92 # Peak to mean
    S_mid = (-0.595 * xiHs + a_S_mid[i]) * b_Hs # Location S-profile
    S_width = (a_S_width[i] * (b_Hs / ((rhos/rhow - 1) * D * xiHs))**0.89) * df; # Width of the S-profile
    S_amp = a_S_amp[i] * df; # Amplitude of the S-profile
    Gx = a_G_x[i]; # Missing element x
    Gz = a_G_z[i]; # Missing element z
    dx = dxx; # dx
    
	# Check the amplitude
    if(df - S_amp < 0):
        print("[Warning] Samp > df, corrected to Samp = df")
        S_amp = df;

    # Convert the width and length of the revetment to a size in which only full STS+ sets will fit
    # STS+ set: width = 1.09 meter, length = 1.2 meter
    m_B = m_Bs * 1.09;
    length = ((d * (zdtop - zdbot))**2 + (slope * d * (zdtop - zdbot))**2)**0.5
    m_Ls = np.ceil(length/1.2)
    length = m_Ls * 1.2
    relzbot = (length**2 / (1 + slope**2))**0.5
    zbot = zdtop * d + d - relzbot
    zdbot = (zbot - d) / d
    
	# Correct the joint filling thickness
    i_fr = i_fr * D;
    

    ########################################
    # Step 4: Load the loading generation function and generate the load
    ########################################
    
    def GenerateLoadingScheme():
        print("[Loading] Started generating the loading")
        # Determine the length of the revetment
        ztop = zdtop * d + d;
        zbot = zdbot * d + d;
        zrev = ztop - zbot;
        revlen = np.sqrt(zrev**2 + (zrev * slope)**2)
        
        # Determine the amount of grid points
        dur = np.ceil(relaxtime / dt)
        for jj in a_Hs:
            dur = dur + np.ceil(impacttime / dt)
        lent = int(dur) 
        lenx = int(np.ceil(revlen / dx)) + 1
        
        # Generate filter thickness array
        bf = np.ones(lenx) * df;
		
		# Do we have a S-profile? If so generatie the variation
        if(S_amp > 0):
            R = 0.03125 * (S_width**2 / S_amp) + 0.5 * S_amp;
            zsmid = d + S_mid
            xsmid = int(round((zsmid - zbot) / (ztop - zbot) * revlen / dx))
            xsbot = int(xsmid - round(0.5 * S_width / dx))
            xstop = int(xsmid + round(0.5 * S_width / dx))
            for i in range(xsbot, xsmid+1):
                sxx = (xsmid - i) * dx
                sxx = sxx - 0.5 * (xsmid - xsbot) * dx
                y = R * np.cos(np.arcsin(sxx / R)) - (R - S_amp)
                bf[i] = bf[i] + y
            for i in range(xsmid, xstop+1):
                sxx = (xstop - i) * dx
                sxx = sxx - 0.5 * (xstop - xsmid) * dx
                y = R * np.cos(np.arcsin(sxx / R)) - (R - S_amp)
                bf[i] = bf[i] - y
        
        
        # Create the grid Ptotal
        Pt = np.zeros((lenx, lent))
        
        # The first x seconds there is no loading due to the relax time
        # After that, we will generate loading wave by wave
        for i in range(len(a_Hs)):
            Hs = a_Hs[i]
            Hs = Hs / 1.4;
            Tm = np.sqrt((2*np.pi*Hs)/(g*s0p))
            
            # Calculate the impact of the wave (Peters, 2017, pg 135-140)
            Lm = (9.81 * Tm**2) / (2.0 * np.pi) # Wave length (assumption: deep water, Tm-1,0 is given)
            eom = (1.0/slope) / (s0p)**0.5 # Iribarren (assumption: Hs = Hm0)
            phimax = (8.0 - 1.6 * eom - 2.0 / ((eom - 0.2)**2)) * Hs # 2% impact
            tr = 0.18 * (phimax / Hs)**-1 # Wave rise time
            td = 4.0 * tr # Wave fall time
            
            # Generate a grid to save the reaction force by this wave
            tr = int(np.ceil(tr / dt))
            td = int(np.ceil(td / dt))
            
			# Define the start and the end of the loading in the grid
            loadingstart = 0
            loadingpeak = int(loadingstart + tr)
            loadingend = int(loadingpeak + td)
            
			# Create a specific grid for this wave impact
            F_t = np.zeros((lenx, loadingend+1))
            F_f = np.zeros((lenx, loadingend+1))
            
            # Impact location on the slope
            impactz = d - 0.7 * Hs
            diffz = impactz - zbot # Difference between bottom revetment and impact location
            impactxs = np.sqrt((diffz)**2 + (slope * diffz)**2) # Length from bottom revetment to impact location
            
            # Create the loading profile in space according to (Peters, 2017)
            # This is the max profile, it only has to be multiplied by the amplitude over time (0 <= amp <= 1)
			# Define the key points
            impactprof = np.zeros(lenx)
            impactmid = int(round(impactxs / dx))                           # Center point of impact
            impacttopdown = int(np.floor((impactxs - (2.0/12.0) * Hs)/dx))      # Down slope top point
            impacttopup = int(np.ceil((impactxs + (2.0/12.0) * Hs)/dx))         # Top slope top point
            impactbottomdown = int(np.floor((impactxs - (7.0/12.0) * Hs)/dx))   # Down slope bottom point
            impactbottomup = int(np.ceil((impactxs + (7.0/12.0) * Hs)/dx))      # Top slope bottom point
            
			# Create the loading profile in space
            warningshown = False;
            warningshown2 = False;
            for x in range(impactbottomdown, impacttopdown+1):
                if(x < 0):
                    x = 0;
                    if(warningshown == False):
                        print("[Loading] Warning: Part of the load falls outside the bottom of the revetment! (H: " + str(Hs) + " [m] & Model: " + str(mname) + ")")
                        warningshown = True;
                a = (float(x) - impactbottomdown) / (impacttopdown - impactbottomdown)
                impactprof[x] = phimax * a;
                
            for x in range(impacttopdown+1, impacttopup):
                if(x < 0):
                    x = 0;
                impactprof[x] = phimax;
            
            for x in range(impacttopup, impactbottomup+1):
                if(x < 0):
                    x = 0;
                if(x < len(impactprof)):
                    a = (impactbottomup - float(x)) / (impactbottomup - impacttopup)
                    impactprof[x] = phimax * a;
                elif(warningshown2 == False):
                    warningshown2 = True;
                    print("[Loading] Warning: Part of the load falls outside the top of the revetment! (H: " + str(Hs) + " [m] & Model: " + str(mname) + ")")
            
            # Create the loading profile in time according to (Peters, 2017)           
            for t in range(loadingstart,loadingpeak+1):
                a = (float(t) - loadingstart) / tr
                F_t[:,t] = a * impactprof;
                
            for t in range(loadingpeak+1, loadingend+1):
                a = (loadingend - float(t)) / td
                F_t[:,t] = a * impactprof;
            
            # Numerical solution to the filter equation with the wave impact profile
            # Method: Implicit Euler
            for t in range(1, loadingend+1):
                size = lenx;
                Mn = np.zeros((size, size))
                
                for j in range(1,size-1):
                    Mn[j, j-1] = 1.0
                    Mn[j, j] = -2.0
                    Mn[j, j+1] = 1.0 - (dx / np.sqrt(D * bf[j] * kk))**2.0
    				
                Mn[0, 0] = -2.0
                Mn[0, 1] = 1.0 - (dx / np.sqrt(D * bf[0] * kk))**2.0
                Mn[-1, -1] = -2.0
                Mn[-1, -2] = 1.0
    
                Ml = - (dx / np.sqrt(D * bf * kk))**2.0 * F_t[:,t-1];
                Mn = np.linalg.inv(Mn)
                F_f[:,t] = np.dot(Mn, Ml)
            
            # Lets assume a force lifting a block from the revetment to be positive
            # The filter pressure is positive and the loading is negative
            time = np.ceil(relaxtime / dt);
            for j in range(0, i):
                time = time + np.ceil(impacttime / dt)
                
            starttime = int(time);
			
			# The total hydraulic head is the filter pressure minus the wave loading
            Pt[:, starttime:starttime+loadingend+1] = F_f - F_t;
        
        # Convert the hydraulic head into pressure
        Pt = Pt * rhow * g; # Pascal = m * kg/m3 * m/s^2
        print("[Loading] Generation succesful")
        return Pt;
    
    
    # Calculate the loading on an element
    # Required input, the loading scheme, the row of the basalton set and the element ID
    def GetElementLoading(P, row, eid):
        # Define the loading
        lent = len(P[0]); # How many time steps do we have?
        t = np.arange(0, lent * dt, dt); # Create an array to store our absolut time steps in
        p = np.zeros(lent); # Create an array to store the results
        
        # Basalton set height
        eid = int(eid)
        setheight = 1.20;
        ehm = (row-1) * setheight + info[eid]['relrot'] # Absolute position of midpoint of the element
        el = info[eid]['length'] # Element length
        ehb = ehm - 0.5 * el
        eht = ehm + 0.5 * el
        
        # Find the neareast grind points for both ends of the basalton element
        ehtg = int(round(eht / dxx))
        ehbg = int(round(ehb / dxx))
        
        # Find the average stress
        for step in range(0, lent):
            p[step] = np.mean(P[ehbg:int(ehtg+1),step])
        
        return p, t;
    
    # This function checks whether a point is within a polygon
	# Matplotlib.pyplot as a nice function to check this, however, it is not available in the Python included in Abaqus
	# Credits: epifanio
	# https://stackoverflow.com/questions/36399381/whats-the-fastest-way-of-checking-if-a-point-is-inside-a-polygon-in-python
    def ray_tracing_method(x,y,poly):
        n = len(poly)
        inside = False
        p1x,p1y = poly[0]
        for i in range(n+1):
            p2x,p2y = poly[i % n]
            if y > min(p1y,p2y):
                if y <= max(p1y,p2y):
                    if x <= max(p1x,p2x):
                        if p1y != p2y:
                            xints = (y-p1y)*(p2x-p1x)/(p2y-p1y)+p1x
                        if p1x == p2x or x <= xints:
                            inside = not inside
            p1x,p1y = p2x,p2y
        return inside
    
    # Generate loading
    if(generateLoading == True):
        P = GenerateLoadingScheme()
        print("Pmax", np.max(P), "N/m2")
        print("Pmin", np.min(P), "N/m2")
    
	
    ########################################
    # Step 5: Generate the bottom of the flume, which is clamped
    ########################################
    
    # Create the part
    mdb.Model(modelType=STANDARD_EXPLICIT, name=str(mname))
    mdb.models[mname].ConstrainedSketch(name='__profile__', sheetSize=10.0)
    mdb.models[mname].sketches['__profile__'].Line(point1=(0.0, 0.0), point2=(m_L, 0.0))
    mdb.models[mname].sketches['__profile__'].HorizontalConstraint(addUndoState=False, entity=mdb.models[mname].sketches['__profile__'].geometry[2])
    mdb.models[mname].Part(dimensionality=THREE_D, name='FlumeBottom', type=ANALYTIC_RIGID_SURFACE)
    mdb.models[mname].parts['FlumeBottom'].AnalyticRigidSurfExtrude(depth=(m_B+1.0), sketch=mdb.models[mname].sketches['__profile__'])
    del mdb.models[mname].sketches['__profile__']
    mdb.models[mname].parts['FlumeBottom'].ReferencePoint(point=mdb.models[mname].parts['FlumeBottom'].vertices[2])
    mdb.models[mname].parts['FlumeBottom'].Surface(name='Surf-FlumeBottom', side1Faces=mdb.models[mname].parts['FlumeBottom'].faces.getSequenceFromMask(('[#1 ]', ), ))
    
    # Place it in the assembly
    mdb.models[mname].rootAssembly.DatumCsysByDefault(CARTESIAN)
    mdb.models[mname].rootAssembly.Instance(dependent=ON, name='FlumeBottom-1', part=mdb.models[mname].parts['FlumeBottom'])
    mdb.models[mname].rootAssembly.translate(instanceList=('FlumeBottom-1', ), vector=(0.0, 0.0, 0.5 * float(m_B)))
    
	
    
    ########################################
    # Step 6: Generate the solid part between the flume bottom and the bottom of the revetment
    # The angle is given by the slope, the thickness is given by b + D
    ########################################
    
    # The bottom part is a solid and represents the concrete structure used during flume experiments
    # It can be described by 4 sets of points and then extruded
    ytr = zdbot * d + d - (t_w**2 / (1 + slope**2))**0.5;   # Top right
    xtr = slope * ytr;                                      # Top right
    xbr = xtr + ((D+b)**2 / (1 + slope**2))**0.5;           # Bottom right
    ybr = ytr - slope * ((D+b)**2 / (1 + slope**2))**0.5;   # Bottom right
    xbl = xbr - slope * ybr;                                # Bottom left
    mdb.models[mname].ConstrainedSketch(name='__profile__', sheetSize=20.0)
    mdb.models[mname].sketches['__profile__'].Line(point1=(0.0, 0.0), point2=(-xtr, ytr))
    mdb.models[mname].sketches['__profile__'].Line(point1=(-xtr, ytr), point2=(-xbr, ybr))
    mdb.models[mname].sketches['__profile__'].PerpendicularConstraint(addUndoState=False, entity1=mdb.models[mname].sketches['__profile__'].geometry[2], entity2=mdb.models[mname].sketches['__profile__'].geometry[3])
    mdb.models[mname].sketches['__profile__'].Line(point1=(-xbr, ybr), point2=(-xbl, 0.0))
    mdb.models[mname].sketches['__profile__'].PerpendicularConstraint(addUndoState=False, entity1=mdb.models[mname].sketches['__profile__'].geometry[3], entity2=mdb.models[mname].sketches['__profile__'].geometry[4])
    mdb.models[mname].sketches['__profile__'].Line(point1=(-xbl, 0.0), point2=(0.0, 0.0))
    mdb.models[mname].sketches['__profile__'].HorizontalConstraint(addUndoState=False, entity=mdb.models[mname].sketches['__profile__'].geometry[5])
    mdb.models[mname].Part(dimensionality=THREE_D, name='FlumeBottomPart', type=ANALYTIC_RIGID_SURFACE)
    mdb.models[mname].parts['FlumeBottomPart'].AnalyticRigidSurfExtrude(depth=(m_B+1.0), sketch=mdb.models[mname].sketches['__profile__'])
    mdb.models[mname].parts['FlumeBottomPart'].Surface(name='Surf-FlumeBottompart', side1Faces=mdb.models[mname].parts['FlumeBottomPart'].faces.getSequenceFromMask(('[#f ]', ), ))
    del mdb.models[mname].sketches['__profile__']
    mdb.models[mname].parts['FlumeBottomPart'].ReferencePoint(point=mdb.models[mname].parts['FlumeBottomPart'].vertices[2])
    
    # Place it in the assembly
    mdb.models[mname].rootAssembly.Instance(dependent=OFF, name='FlumeBottomPart-1', part=mdb.models[mname].parts['FlumeBottomPart'])
    mdb.models[mname].rootAssembly.translate(instanceList=('FlumeBottomPart-1', ), vector=(0.0, 0.0, 0.5 * float(m_B)))
    
    mdb.models[mname].rootAssembly.DatumCsysByThreePoints(coordSysType=
        CARTESIAN, line2=(0.54355, 0.54355, 0.0), name='Datum csys-4', origin=
        mdb.models[mname].rootAssembly.instances['FlumeBottomPart-1'].vertices[5]
        , point1=
        mdb.models[mname].rootAssembly.instances['FlumeBottomPart-1'].InterestingPoint(
        mdb.models[mname].rootAssembly.instances['FlumeBottomPart-1'].edges[9], 
        MIDDLE))
	
    
	
    ########################################
    # Step 7: Generate the bottom of the model underneath the revetment
    # Take into account the S-profile!
    ########################################
    
	# Old coordinates for a part which is no longer used, however those coordinates are still required to generate this part of the model
	yt = zdtop * d + d;
    xt = slope * yt;
    xb = xt + ((D+b)**2 / (1 + slope**2))**0.5;
    yb = yt - slope * ((D+b)**2 / (1 + slope**2))**0.5;
	
    # Coordinates to create S-profile	
    y = d + S_mid * revHs; # Zsmid on top layer
    x = y * slope; # Zsmid on top layer
    xgm = x + ((D+b)**2 / (1 + slope**2))**0.5; # Zsmid underneath filter
    ygm = y - slope * ((D+b)**2 / (1 + slope**2))**0.5; # Zsmid underneath filter
    xgb = xgm - slope * ((0.5 * S_width)**2 / (1 + slope**2))**0.5; # Zsbot underneath filter (0.5 Bs)
    ygb = ygm - ((0.5 * S_width)**2 / (1 + slope**2))**0.5; # Zsbot underneath filter (0.5 Bs)
    xgt = xgm + slope * ((0.5 * S_width)**2 / (1 + slope**2))**0.5; # Zstop underneath filter (0.5 Bs)
    ygt = ygm + ((0.5 * S_width)**2 / (1 + slope**2))**0.5; # Zstop underneath filter (0.5 Bs)
    xt = xb; # Ztopdom underneath filter
    yt = yb; # Ztopdom underneath filter
    xb = xbr; # Zbotdom underneath filter
    yb = ybr; # Zbotdom underneath filter
    xhump = 0.5 * (xgb + xgm) - ((S_amp)**2 / (1 + slope**2))**0.5; # Third point of the arc
    yhump = 0.5 * (ygb + ygm) + slope * ((S_amp)**2 / (1 + slope**2))**0.5; # Third point of the arc
    xtrough = 0.5 * (xgt + xgm) + ((S_amp)**2 / (1 + slope**2))**0.5; # Third point of the arc
    ytrough = 0.5 * (ygt + ygm) - slope * ((S_amp)**2 / (1 + slope**2))**0.5; # Third point of the arc
    
    # Create part
    mdb.models[mname].ConstrainedSketch(name='__profile__', sheetSize=50.0)
    if(S_amp > 0):
        # If there is a S-profile...
        mdb.models[mname].sketches['__profile__'].Line(point1=(-xb, yb), point2=(-xgb, ygb))
        mdb.models[mname].sketches['__profile__'].Arc3Points(point1=(-xgb, ygb), point2=(-xgm, ygm), point3=(-xhump, yhump))
        mdb.models[mname].sketches['__profile__'].Arc3Points(point1=(-xgm, ygm), point2=(-xgt, ygt), point3=(-xtrough, ytrough))
        mdb.models[mname].sketches['__profile__'].Line(point1=(-xgt, ygt), point2=(-xt, yt))
    else:
        # Otherwise no curve
        mdb.models[mname].sketches['__profile__'].Line(point1=(-xb, yb), point2=(-xt, yt))
    mdb.models[mname].Part(dimensionality=THREE_D, name='FlumeCore', type=ANALYTIC_RIGID_SURFACE)
    mdb.models[mname].parts['FlumeCore'].AnalyticRigidSurfExtrude(depth=(m_B+1.0), sketch=mdb.models[mname].sketches['__profile__'])
    mdb.models[mname].parts['FlumeCore'].Surface(name='Surf-FlumeCore', side2Faces=mdb.models[mname].parts['FlumeCore'].faces.getSequenceFromMask(('[#1 ]', ), ))
    del mdb.models[mname].sketches['__profile__']
    mdb.models[mname].parts['FlumeCore'].ReferencePoint(point=mdb.models[mname].parts['FlumeCore'].vertices[2])
    
    # Place it in the assembly
    mdb.models[mname].rootAssembly.DatumCsysByDefault(CARTESIAN)
    mdb.models[mname].rootAssembly.Instance(dependent=ON, name='FlumeCore-1', part=mdb.models[mname].parts['FlumeCore'])
    mdb.models[mname].rootAssembly.translate(instanceList=('FlumeCore-1', ), vector=(0.0, 0.0, 0.5 * float(m_B)))

    
    
    ########################################
    # Step 8: Generate the top layer
    # Take into account 'removed' elements!
    ########################################
    
	# We need to create a surface later, the only way I know how to do this is to hardcode the surface IDs
    surfid = ['[#3fff ]', '[#3ff ]', '[#fff ]', '[#7ff ]', '[#1fff ]', '[#fff ]', '[#3fff ]', '[#fff ]', '[#7fff ]', '[#7fff ]', '[#1ff ]', '[#3fff ]', '[#fff ]', '[#7ff ]', '[#7ff ]', '[#3fff ]', '[#1fff ]', '[#fff ]']
    
	# Loop through element 1 to 18 and create the parts
	for bid in range(1, 19):
        bname = 'block' + str(bid);
        mdb.models[mname].ConstrainedSketch(name='__profile__', sheetSize=50.0)
        blockxy = np.loadtxt('H:\Desktop\Model\Basalton\STS block' + str(bid) + '.txt', delimiter=',')
        for i in np.arange(len(blockxy)-1):
            x1 = blockxy[i][0];
            y1 = blockxy[i][1];
            x2 = blockxy[i+1][0];
            y2 = blockxy[i+1][1];
            mdb.models[mname].sketches['__profile__'].Line(point1=(x1, y1), point2=(x2, y2))
        mdb.models[mname].Part(dimensionality=THREE_D, name=bname, type=DEFORMABLE_BODY)
        mdb.models[mname].parts[bname].BaseSolidExtrude(depth=D, sketch=mdb.models[mname].sketches['__profile__'])
        del mdb.models[mname].sketches['__profile__']
        
        mdb.models[mname].parts[bname].seedPart(deviationFactor=0.1, minSizeFactor=0.1, size=10.0)
        mdb.models[mname].parts[bname].generateMesh()
        
        # Create a surface
        mdb.models[mname].parts['block' + str(bid)].Surface(name='surface', side1Faces=mdb.models[mname].parts['block' + str(bid)].faces.getSequenceFromMask((surfid[bid-1], ), ))
        
		# Place it in the assembly
        insn = 'block' + str(bid);
        mdb.models[mname].rootAssembly.Instance(dependent=ON, name=insn, part=mdb.models[mname].parts[bname])
        mdb.models[mname].rootAssembly.rotate(angle=-90.0, axisDirection=(1.0, 0.0, 0.0), axisPoint=(0.0, 0.0, 0.0), instanceList=(insn, ))
        mdb.models[mname].rootAssembly.rotate(angle=90.0, axisDirection=(0.0, 1.0, 0.0), axisPoint=(0.0, 0.0, 0.0), instanceList=(insn, ))
        mdb.models[mname].rootAssembly.rotate(angle=-(np.arctan(1/slope)*360)/(2*np.pi), axisDirection=(0.0, 0.0, 1.0), axisPoint=(0.0, 0.0, 0.0), instanceList=(insn, ))
        mdb.models[mname].rootAssembly.translate(instanceList=(insn, ), vector=(0.0, 0.0, m_B))
        mdb.models[mname].rootAssembly.translate(instanceList=(insn, ), vector=(-np.sqrt(D**2*(1+slope**2)), 0.0, 0.0))
        mx = (zdbot * d + d) * slope - slope**2 * np.sqrt(D**2 / (1+slope**2));
        my = zdbot * d + d - slope * np.sqrt(D**2 / (1+slope**2));
        mdb.models[mname].rootAssembly.translate(instanceList=(insn, ), vector=(-mx, my, 0))
        
	# Copy paste the sets the required number of times
    mdb.models[mname].rootAssembly.LinearInstancePattern(direction1=(0.0, 0.0, -1.0), direction2=(-1, 1/slope, 0.0), instanceList=('block1', 
        'block2', 'block3', 'block4', 'block5', 'block6', 'block7', 
        'block8', 'block11', 'block13', 'block14', 'block12', 
        'block18', 'block17', 'block16', 'block10', 'block9', 
        'block15'), number1=int(m_Bs), number2=int(m_Ls), spacing1=1.09, spacing2=1.2)
    
    # Rename the blocks to the new naming convention
    # block1-lin-2-5 -> Element 1 on column 2 (from the right) and row 5 (from the bottom)
    for bid in range(1,19):
        mdb.models[mname].rootAssembly.features.changeKey(fromName='block' + str(bid), toName='block'  + str(bid) + '-lin-1-1')
    
    # Generate a new material, Basalton
    mdb.models[mname].Material(name='Basalton')
    mdb.models[mname].materials['Basalton'].Density(table=((rhos - rhow, ), ))  # Submerged density
    mdb.models[mname].materials['Basalton'].Elastic(table=((50000000000.0, 0.2), ))  # Concrete
    mdb.models[mname].HomogeneousSolidSection(material='Basalton', name='Sec-Basalton', thickness=None)
    
    # Add Basalton to all elements
    for bid in range(1,19):
        partname = 'block' + str(bid);
        mdb.models[mname].parts[partname].Set(cells=mdb.models[mname].parts[partname].cells.getSequenceFromMask(('[#1 ]', ), ), name='Set-1')
        mdb.models[mname].parts[partname].SectionAssignment(offset=0.0, offsetField='', offsetType=MIDDLE_SURFACE, region=mdb.models[mname].parts[partname].sets['Set-1'], sectionName='Sec-Basalton', thicknessAssignment=FROM_SECTION)
        mdb.models[mname].rootAssembly.regenerate()
    
    delblock = 'block0-lin-0-0'
	# Do we need to remove an element?
    if(Gz != 0):
        # We need to remove an element
		
		# Project x and z on the slope
        Gztop = zdtop * d + d
        Growheight = np.sqrt(1.2**2 / (1 + slope**2))
        Gz = Gz - (zdbot * d + d)
        Grow = np.ceil(Gz / Growheight)
        Gcolumn = 2 + np.floor((Gx + 1.09/2)/ 1.09)
        relGx = (1.5 * 1.09 + Gx) - (Gcolumn - 1.0) * 1.09
        relGzs = np.sqrt(slope**2 * Gz**2 + Gz**2)
        relGzs = relGzs - (Grow - 1.0) * 1.2
        print("Rel X and Z", relGx, relGzs)
        
		# Find out which elements is at (x, z)
        blockid = -1;
        for bid in range(1,19):
            blockxy = np.loadtxt('H:\Desktop\Model\Basalton\STS block' + str(bid) + '.txt', delimiter=',')
            inside = ray_tracing_method(relGx, relGzs, blockxy)
            if(inside == True):
                blockid = bid;
                break;
                
		# The point (x, z) is an open space, select the element nearest to the point
        if(blockid == -1):
            print("Sample points falls in open space, selecting the element most nearby")
            Gmindis = 999;
            GmindisID = -1
            for bid in range(1,19):
                blockxy = np.loadtxt('H:\Desktop\Model\Basalton\STS block' + str(bid) + '.txt', delimiter=',')
                for points in blockxy:
                    Gminx = points[0] #(np.max(blockxy[0]) - np.min(blockxy[0])) * (1/2) + np.min(blockxy[0])
                    Gminz = points[1] #(np.max(blockxy[1]) - np.min(blockxy[1])) * (1/2) + np.min(blockxy[1])
                    Gdeltax = np.abs(relGx - Gminx)
                    Gdeltaz = np.abs(relGzs - Gminz)
                    Gdis = np.sqrt(Gdeltax**2 + Gdeltaz**2)
                    if(Gdis < Gmindis):
                        Gmindis = Gdis
                        GmindisID = bid
            blockid = GmindisID
                
		# Delete the corresponding block from the assembly
        print("Delete", blockid, "in row", Grow, "and column", Gcolumn)
        delblock = 'block' + str(blockid) + '-lin-' + str(int(Gcolumn)) + '-' + str(int(Grow))
        del mdb.models[mname].rootAssembly.features[delblock]
    
    
    
    ########################################
    # Constraints
    ########################################
    
	# Elements at the border
    botele = [15, 16, 17, 18]
    leftele = [1, 5, 9, 15]
    topele = [1, 2, 3, 4]
    rightele = [4, 8, 13, 14, 18]
    
	# Apply ZSYMM boundary conditions to the elements in the most right and most left column of the revetment
    count = 1
    for row in range(1, int(m_Ls)+1):
        for eid in leftele:
            # Left
            bname = 'block' + str(eid) + '-lin-1-' + str(int(row));
            mdb.models[mname].rootAssembly.Set(cells=mdb.models[mname].rootAssembly.instances[bname].cells.getSequenceFromMask(('[#1 ]', ), ), name='Set-' + str(count))
            mdb.models[mname].ZsymmBC(createStepName='Initial', localCsys=mdb.models[mname].rootAssembly.datums[6], name='BC-' + str(count), region=mdb.models[mname].rootAssembly.sets['Set-' + str(count)])
            count = count + 1;
            
        for eid in rightele:
            # Right
            bname = 'block' + str(eid) + '-lin-' + str(int(m_Bs)) + '-' + str(int(row));
            mdb.models[mname].rootAssembly.Set(cells=mdb.models[mname].rootAssembly.instances[bname].cells.getSequenceFromMask(('[#1 ]', ), ), name='Set-' + str(count))
            mdb.models[mname].ZsymmBC(createStepName='Initial', localCsys=mdb.models[mname].rootAssembly.datums[6], name='BC-' + str(count), region=mdb.models[mname].rootAssembly.sets['Set-' + str(count)])
            count = count + 1;
    
	# Define the block - filter interaction
    # Contact Property
    mdb.models[mname].ContactProperty('Block-Filter')
    mdb.models[mname].interactionProperties['Block-Filter'].TangentialBehavior(dependencies=0, directionality=ISOTROPIC, elasticSlipStiffness=None, formulation=PENALTY, fraction=0.005, maximumElasticSlip=FRACTION, pressureDependency=OFF, shearStressLimit=None, slipRateDependency=OFF, table=((Fblockfilter, ), ), temperatureDependency=OFF)
    mdb.models[mname].interactionProperties['Block-Filter'].NormalBehavior(allowSeparation=ON, constraintEnforcementMethod=DEFAULT, pressureOverclosure=HARD)
    
    # Add the defined interaction with the filter to each element
    mdb.models[mname].rootAssembly.Surface(name='s_ondergrond', side2Faces=mdb.models[mname].rootAssembly.instances['FlumeCore-1'].faces.getSequenceFromMask( ('[#1 ]', ), ))
    for column in range(1, int(m_Bs)+1):
        for row in range(1, int(m_Ls)+1):
            for bid in range(1, 19):
                bname = 'block' + str(bid) + '-lin-' + str(column) + '-' + str(row);
                if(bname != delblock):
                    mdb.models[mname].SurfaceToSurfaceContactExp(clearanceRegion=None, createStepName='Initial', datumAxis=None, initialClearance=OMIT, interactionProperty='Block-Filter', master=mdb.models[mname].rootAssembly.instances[bname].surfaces['surface'], mechanicalConstraint=PENALTY, name='Int-' + str(bname), slave=mdb.models[mname].rootAssembly.surfaces['s_ondergrond'], sliding=FINITE)
                
                
    
    ########################################
    # Step 9: Generate the joint filling
    # This part is not used within this study to cut down on computation time
    ########################################
    
    # Allow also for tests without joint filling
    if(i_fr > 0):
		# Loop through all joints
        for iid in range(1, 25):
			# Create a part for the joint filling
            iname = 'filling' + str(iid);
            mdb.models[mname].ConstrainedSketch(name='__profile__', sheetSize=50.0)
            blockxy = np.loadtxt('H:\Desktop\Model\Inwas\Inwas' + str(iid) + '.txt', delimiter=',')
            for i in np.arange(len(blockxy)-1):
                x1 = blockxy[i][0];
                y1 = blockxy[i][1];
                x2 = blockxy[i+1][0];
                y2 = blockxy[i+1][1];
                mdb.models[mname].sketches['__profile__'].Line(point1=(x1, y1), point2=(x2, y2))
            mdb.models[mname].Part(dimensionality=THREE_D, name=iname, type=DEFORMABLE_BODY)
            mdb.models[mname].parts[iname].BaseSolidExtrude(depth=i_fr, sketch=mdb.models[mname].sketches['__profile__'])
            del mdb.models[mname].sketches['__profile__']
            
			# Seed the part
            mdb.models[mname].parts[iname].seedPart(deviationFactor=0.1, minSizeFactor=0.1, size=1)
            mdb.models[mname].parts[iname].setMeshControls(elemShape=WEDGE, regions=mdb.models[mname].parts[iname].cells.getSequenceFromMask(('[#1 ]', ), ))
            mdb.models[mname].parts[iname].generateMesh()
        
			# Add the joint filling to the assembly
            mdb.models[mname].rootAssembly.Instance(dependent=ON, name=iname, part=mdb.models[mname].parts[iname])
            mdb.models[mname].rootAssembly.rotate(angle=-90.0, axisDirection=(1.0, 0.0, 0.0), axisPoint=(0.0, 0.0, 0.0), instanceList=(iname, ))
            mdb.models[mname].rootAssembly.rotate(angle=90.0, axisDirection=(0.0, 1.0, 0.0), axisPoint=(0.0, 0.0, 0.0), instanceList=(iname, ))
            mdb.models[mname].rootAssembly.rotate(angle=-(np.arctan(1/slope)*360)/(2*np.pi), axisDirection=(0.0, 0.0, 1.0), axisPoint=(0.0, 0.0, 0.0), instanceList=(iname, ))
            mdb.models[mname].rootAssembly.translate(instanceList=(iname, ), vector=(0.0, 0.0, m_B))
            mdb.models[mname].rootAssembly.translate(instanceList=(iname, ), vector=(-np.sqrt(D**2*(1+slope**2)), 0.0, 0.0))
            mx = (zdbot * d + d) * slope - slope**2 * np.sqrt(D**2 / (1+slope**2));
            my = zdbot * d + d - slope * np.sqrt(D**2 / (1+slope**2));
            mdb.models[mname].rootAssembly.translate(instanceList=(iname, ), vector=(-mx, my, 0))
            
		# Copy paste the joint filling the required amount of times
        mdb.models[mname].rootAssembly.LinearInstancePattern(direction1=(0.0, 0.0, -1.0), direction2=(-1, 1/slope, 0.0), instanceList=('filling1', 'filling2', 'filling3', 'filling4', 'filling5', 
            'filling6', 'filling7', 'filling8', 'filling9', 'filling10', 'filling11', 'filling12', 'filling13', 'filling14', 'filling15', 'filling16', 'filling17', 'filling18', 'filling19', 'filling20',
            'filling21', 'filling22', 'filling23', 'filling24'), number1=int(m_Bs), number2=int(m_Ls), spacing1=1.09, spacing2=1.2)
        
        # Rename the blocks to the new naming convention
        # filling-lin-2-5 -> Filling 1 on column 2 (from the right) and row 5 (from the bottom)
        for iid in range(1,25):
            mdb.models[mname].rootAssembly.features.changeKey(fromName='filling' + str(iid), toName='filling'  + str(iid) + '-lin-1-1')
            
        # Generate a new material, Joint Filling
        mdb.models[mname].Material(name='Filling')
        mdb.models[mname].materials['Filling'].Density(table=(((rhoi - rhow) * (1 - ni), ), ))
        mdb.models[mname].materials['Filling'].Elastic(table=((70000000000.0, 0.3), )) # Granite
        mdb.models[mname].HomogeneousSolidSection(material='Filling', name='Sec-Filling', thickness=None)
        
        # Add Joint filling to all elements
        for bid in range(1,25):
            partname = 'filling' + str(bid);
            mdb.models[mname].parts[partname].Set(cells=mdb.models[mname].parts[partname].cells.getSequenceFromMask(('[#1 ]', ), ), name='Set-1')
            mdb.models[mname].parts[partname].SectionAssignment(offset=0.0, offsetField='', offsetType=MIDDLE_SURFACE, region=mdb.models[mname].parts[partname].sets['Set-1'], sectionName='Sec-Filling', thicknessAssignment=FROM_SECTION)
            mdb.models[mname].rootAssembly.regenerate()
            
        # Delete the joint filling which are not fully enclosed by elements (at the borders)
        delrow = [15, 16, 17, 18, 19];
        delcol = [19, 20, 21, 22, 23, 24];
        delcombi = [15, 16, 17, 18, 19, 20, 21, 22, 23, 24];
        
        # At the right side
        if(m_Bs > 1):
            row = int(m_Ls);
            for column in range(1, m_Bs):
                for fillid in delrow:
                    bname = 'filling' + str(fillid) + '-lin-' + str(column) + '-' + str(row);
                    del mdb.models[mname].rootAssembly.features[bname]
                    
        # At the top
		if(m_Ls > 1):
            column = int(m_Bs);
            for row in range(1, int(m_Ls)):
                for fillid in delcol:
                    bname = 'filling' + str(fillid) + '-lin-' + str(column) + '-' + str(row);
                    del mdb.models[mname].rootAssembly.features[bname]
        
		# Corners
        for fillid in delcombi:
            column = int(m_Bs)
            row = int(m_Ls)
            bname = 'filling' + str(fillid) + '-lin-' + str(column) + '-' + str(row);
            del mdb.models[mname].rootAssembly.features[bname]
    


    ########################################
    # Step 9: S-profile correction
    # Create a S-profile on elements and filling
    ########################################
    
    # Return the slope coordinate (0 = zbot, going up to ztop)
    def GetSlopeCoord(z):
        zbot = zdbot * d + d;
        sc = ((z-zbot)**2 + ((z-zbot)*slope)**2)**0.5
        if((z - zbot) < 0):
            sc = 0;
        return sc;
    
    def GetSlopeCoordBlock(row, block):
        base = (row-1) * 1.2;
        sc = base + info[block]['relrot']
        return sc;
    
    def GetSlopeCoordInwas(row, inwas):
        base = (row-1) * 1.2;
        sc = base + inwasinfo[inwas]['relrot']
        return sc;
    
    
    # Loop through every element, check the slope coord and adjust for a S-profile
    # block1-lin-2-5 -> Element 1 on column 2 (from the right) and row 5 (from the bottom)
    if(S_amp > 0):
        smid = GetSlopeCoord(d+S_mid*revHs)
        sbot = smid - 0.5 * S_width
        stop = smid + 0.5 * S_width
        
		# Radius of the S-profile
        R = 0.03125 * (S_width**2 / S_amp) + 0.5 * S_amp;
        
		# Check if we also need to adjust the joint filling in height
        maxc = 43;
        if(i_fr == 0.0):
            maxc = 19;
		
		# Loop through all elements (and if enabled, joint filling) and correct their position
        for row in np.arange(1, 1+m_Ls):
            for column in np.arange(1, 1+m_Bs):
                for block in np.arange(1, maxc):
                    zakking = 0;
                    
                    if(block < 19):
                        sx = GetSlopeCoordBlock(row, block);
                        sxt = sx + 0.5 * info[block]['length'];
                        sxb = sx - 0.5 * info[block]['length'];
                        bname = 'block' + str(block) + '-lin-' + str(column) + '-' + str(int(row));
                    else:
                        block = block - 18
                        sx = GetSlopeCoordInwas(row, block);
                        sxt = sx + 0.5 * inwasinfo[block]['length'];
                        sxb = sx - 0.5 * inwasinfo[block]['length'];
                        bname = 'filling' + str(block) + '-lin-' + str(column) + '-' + str(int(row));
                    
                    if(sxt > sbot and sxt <= 0.5 * (sbot + smid)):
                        # Below hump
                        # Align height with the upper corner of the elements
                        bx = 0.25 * S_width - (sxt - sbot);
                        zakking = np.sin(np.arccos(bx/R)) * R - (R - S_amp)
                        
                    elif(sxt > 0.5 * (sbot+smid) and sxb < 0.5 * (sbot + smid)):
                        # On top of the hump
                        zakking = S_amp;
                        
                    elif(sxb >= 0.5 * (sbot + smid) and sx <= 0.5 * (smid + stop)):
                        # Between hump and trough, align on the lower corner of the elements
                        if(sx <= smid):
                            bx = 0.25 * S_width - (sxb - sbot)
                            zakking = np.sin(np.arccos(bx/R)) * R - (R - S_amp)
                        else:
                            bx = 0.25 * S_width - (sxb - smid)
                            zakking = np.sin(np.arccos(bx/R)) * R - (R - S_amp)
                            zakking = zakking * -1;
                        
                    elif(sx > 0.5 * (smid + stop) and sxt < stop):
                        # Between trough and normal revetment, align on the upper corner
                        bx = 0.25 * S_width - (sxt - smid)
                        zakking = np.sin(np.arccos(bx/R)) * R - (R - S_amp)
                        zakking = zakking * -1;            
                    
                    if(zakking != 0):
                        angle = np.arctan(slope);
                        dx = np.cos(angle) * zakking;
                        dy = np.sin(angle) * zakking;
                        mdb.models[mname].rootAssembly.translate(instanceList=(bname, ), vector=(dx, dy, 0.0))
                        
                        
    
    ########################################
    # Step 10: Setup the model
    ########################################
    
	# Set the correct simulation time
    totaltime = relaxtime
    if(generateLoading == True):
        totaltime = totaltime + impacttime * len(a_Hs);
        
    mdb.models[mname].ExplicitDynamicsStep(description='Idle time', improvedDtMethod=ON, maxIncrement=0.1, name='Idle', previous='Initial', timePeriod=totaltime)
    mdb.models[mname].Gravity(comp2=-g, createStepName='Idle', distributionType=UNIFORM, field='', name='Gravity')
    
    # Encastre the ground surfaces
    mdb.models[mname].rootAssembly.Set(name='Set-91', referencePoints=((
        mdb.models[mname].rootAssembly.instances['FlumeCore-1'].referencePoints[3], ), (
        mdb.models[mname].rootAssembly.instances['FlumeBottomPart-1'].referencePoints[3], ), (
        mdb.models[mname].rootAssembly.instances['FlumeBottom-1'].referencePoints[2], )))
    mdb.models[mname].EncastreBC(createStepName='Idle', localCsys=None, name='BC-91', region=mdb.models[mname].rootAssembly.sets['Set-91'])
    
	# Model the global default contact (block-block)
    mdb.models[mname].ContactProperty('Block-Block')
    mdb.models[mname].interactionProperties['Block-Block'].TangentialBehavior(dependencies=0, directionality=ISOTROPIC, elasticSlipStiffness=None, formulation=PENALTY, fraction=0.005, maximumElasticSlip=FRACTION, pressureDependency=OFF, shearStressLimit=None, slipRateDependency=OFF, table=((Fblockblock, ), ), temperatureDependency=OFF)
    mdb.models[mname].interactionProperties['Block-Block'].NormalBehavior(allowSeparation=ON, constraintEnforcementMethod=DEFAULT, pressureOverclosure=HARD)
    
    mdb.models[mname].ContactExp(createStepName='Idle', name='Block-BlockInt')
    mdb.models[mname].interactions['Block-BlockInt'].includedPairs.setValuesInStep(stepName='Idle', useAllstar=ON)
    mdb.models[mname].interactions['Block-BlockInt'].contactPropertyAssignments.appendInStep(assignments=((GLOBAL, SELF, 'Block-Block'), ), stepName='Idle')
    
	# Add the loading to the elements
    if(generateLoading == True):
        # On which surface do we want to apply the load? Unfortunately this has to be hardcoded as Abaqus does not support selection of surfaces through a function
        surfaceID = np.array(['[#1000 ]', '[#100 ]', '[#400 ]', '[#200 ]', '[#800 ]', '[#400 ]', '[#1000 ]', '[#400 ]', '[#2000 ]', '[#2000 ]', '[#80 ]', '[#1000 ]', '[#400 ]', '[#200 ]', '[#200 ]', '[#1000 ]', '[#800 ]', '[#400 ]']);
        
		# Loop through all elements
        for row in range(1, int(m_Ls)+1):
            for bid in range(1, 19):
                pel, tel = GetElementLoading(P, row, bid)
                loading = np.array((tel, pel))
                loading = np.transpose(loading)
                loading = map(tuple, loading)
                loading = tuple(loading)
                ampname = 'Amp' + str(bid) + '-lin-' + str(row);
                mdb.models[mname].TabularAmplitude(data=(loading), name=ampname, smooth=SOLVER_DEFAULT, timeSpan=STEP)
                
                for column in range(1, m_Bs+1):
                    surfname = 'Surf' + str(bid) + '-lin-' + str(column) + '-' + str(row);
                    loadname = 'Load' + str(bid) + '-lin-' + str(column) + '-' + str(row);
                    blockname = 'block' + str(bid) + '-lin-' + str(column) + '-' + str(row);
                    if(blockname != delblock):
                        mdb.models[mname].rootAssembly.Surface(name=surfname, side1Faces=mdb.models[mname].rootAssembly.instances[blockname].faces.getSequenceFromMask((surfaceID[bid-1], ), ))
                        mdb.models[mname].Pressure(amplitude=UNSET, createStepName='Idle', distributionType=UNIFORM, field='', magnitude=-1.0, name=loadname, region=mdb.models[mname].rootAssembly.surfaces[surfname])
                        mdb.models[mname].loads[loadname].setValues(amplitude=ampname)
    
    # Request field data
    mdb.models[mname].fieldOutputRequests['F-Output-1'].setValues(numIntervals=int(totaltime/dt))
    mdb.models[mname].fieldOutputRequests['F-Output-1'].setValues(variables=('U', ))
    mdb.models[mname].historyOutputRequests['H-Output-1'].setValues(variables=('DT', ))

	# Generate an .inp file
    jname = 'Job-' + str(mname);
    mdb.jobs[jname].writeInput(consistencyChecking=OFF)
    print("Model " + str(mname) + " finished in " + str(round(time.time() - modelstart, 2)) + " sec | totale tijd: " + str(round(time.time() - start, 2)) + " sec");
   
# Delete the empty model   
del mdb.models['Model-1']